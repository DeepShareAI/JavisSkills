#!/usr/bin/env python3
"""Read an .xlsx file and emit JSON.

Usage:
    read_xlsx.py <path> [--sheet NAME] [--range A1:D20]

With --sheet: emits [{header: value, ...}, ...] for that sheet.
Without --sheet: emits {sheet_name: [...rows], ...} for all sheets.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _lib


def _to_rows(ws, range_: str | None) -> list[dict[str, Any]]:
    if range_:
        cells = ws[range_]
        rows = [[c.value for c in row] for row in cells]
    else:
        rows = [[c.value for c in row] for row in ws.iter_rows()]
    if not rows:
        return []
    header = [v if v is not None else f"col{i}" for i, v in enumerate(rows[0])]
    return [dict(zip(header, r)) for r in rows[1:]]


def run(argv: list[str]) -> None:
    parser = argparse.ArgumentParser(prog="read_xlsx.py")
    parser.add_argument("path")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--range", dest="range_", default=None)
    args = parser.parse_args(argv)

    try:
        target = _lib.resolve_under_root(args.path)
    except _lib.TraversalError as e:
        _lib.emit_err(str(e), code=2)

    if not target.exists():
        _lib.emit_err(f"path not found: '{args.path}'", code=4)
    if not target.is_file():
        _lib.emit_err(f"not a regular file: '{args.path}'", code=5)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(target, data_only=True)
    except Exception as e:
        _lib.emit_err(f"failed to parse '{args.path}': {e}", code=6)

    if args.sheet is not None:
        if args.sheet not in wb.sheetnames:
            _lib.emit_err(f"sheet '{args.sheet}' not in workbook", code=7)
        _lib.emit_ok(_to_rows(wb[args.sheet], args.range_))
        return

    _lib.emit_ok({name: _to_rows(wb[name], args.range_) for name in wb.sheetnames})


if __name__ == "__main__":
    run(sys.argv[1:])
